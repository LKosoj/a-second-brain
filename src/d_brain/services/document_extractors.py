"""Document extraction helpers for uploaded Telegram files."""

from __future__ import annotations

import html
import re
import subprocess
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from d_brain.services.json_normalizer import extract_first_json_dict

SUPPORTED_DOCUMENT_FORMATS = frozenset(
    {"pdf", "docx", "xlsx", "txt", "md", "html", "pptx", "mpp"}
)
TEXT_LIKE_DOCUMENT_FORMATS = frozenset({"txt", "md", "html"})
DOCUMENT_DOWNLOAD_MAX_BYTES = 20 * 1024 * 1024
DOCUMENT_MAX_PDF_PAGES = 1000
DOCUMENT_MAX_PPTX_SLIDES = 80
DOCUMENT_MAX_XLSX_SHEETS = 20
DOCUMENT_EXTRACTION_TIMEOUT_SECONDS = 600
MPP_READER_TIMEOUT_SECONDS = max(10, DOCUMENT_EXTRACTION_TIMEOUT_SECONDS - 5)
_REPO_ROOT = Path(__file__).resolve().parents[3]
_MPP_READER_SCRIPT_CANDIDATES = (
    _REPO_ROOT
    / "vault"
    / ".claude"
    / "skills"
    / "ms-project"
    / "scripts"
    / "mpp_read.py",
    Path.home() / ".codex" / "skills" / "ms-project" / "scripts" / "mpp_read.py",
)

_MIME_TO_FORMAT = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": "pptx",
    "application/vnd.ms-project": "mpp",
    "text/plain": "txt",
    "text/markdown": "md",
    "text/x-markdown": "md",
    "text/html": "html",
    "application/xhtml+xml": "html",
}


def detect_document_format(file_name: str, mime_type: str = "") -> str | None:
    """Resolve one supported format from filename and Telegram mime metadata."""

    suffix = Path(file_name or "").suffix.lower().lstrip(".")
    if suffix == "htm":
        suffix = "html"
    if suffix in SUPPORTED_DOCUMENT_FORMATS:
        return suffix
    normalized_mime = (mime_type or "").strip().lower()
    return _MIME_TO_FORMAT.get(normalized_mime)


def extract_document_payload(
    input_path: str | Path,
    *,
    file_format: str,
    original_name: str,
) -> dict[str, Any]:
    """Extract one saved document into the shared normalized payload."""

    path = Path(input_path)
    resolved_format = "html" if file_format == "htm" else file_format
    warnings: list[str] = []
    metadata: dict[str, Any] = {
        "original_name": original_name,
        "size_bytes": path.stat().st_size,
    }
    truncated = False
    title = _title_from_name(original_name)
    plain_text = ""

    if resolved_format not in SUPPORTED_DOCUMENT_FORMATS:
        warnings.append(f"Unsupported format: {resolved_format}")
    elif resolved_format in TEXT_LIKE_DOCUMENT_FORMATS:
        plain_text, title, extra_metadata, extra_warnings = _extract_text_like(
            path,
            resolved_format,
            original_name=original_name,
        )
        metadata.update(extra_metadata)
        warnings.extend(extra_warnings)
    else:
        plain_text, title, extra_metadata, extra_warnings = _extract_binary_document(
            path,
            resolved_format,
            original_name=original_name,
        )
        metadata.update(extra_metadata)
        warnings.extend(extra_warnings)

    normalized_text = _normalize_plain_text(plain_text)
    if not normalized_text:
        warnings.append("No extracted text available.")

    return {
        "plain_text": normalized_text,
        "title": title or _title_from_name(original_name),
        "format": resolved_format,
        "warnings": _unique_ordered(warnings),
        "metadata": metadata,
        "truncated": truncated,
    }


def summarize_plain_text(text: str, *, limit: int) -> str:
    """Build one deterministic short summary from extracted plain text."""

    compact = re.sub(r"\s+", " ", text or "").strip()
    if limit > 0 and len(compact) > limit:
        return f"{compact[:limit].rstrip()}..."
    return compact


def _extract_text_like(
    path: Path,
    file_format: str,
    *,
    original_name: str,
) -> tuple[str, str, dict[str, Any], list[str]]:
    data = path.read_bytes()
    decoded, encoding, warnings = _decode_text_bytes(data)
    metadata: dict[str, Any] = {
        "encoding": encoding,
        "characters": len(decoded),
    }

    if file_format == "html":
        extracted = _html_to_text(decoded)
        title = extracted["title"] or _title_from_name(original_name)
        plain_text = extracted["content"]
    elif file_format == "md":
        title = _markdown_title(decoded) or _title_from_name(original_name)
        plain_text = _markdown_to_plain_text(decoded)
    else:
        title = _title_from_name(original_name)
        plain_text = decoded

    return plain_text, title, metadata, warnings


def _extract_binary_document(
    path: Path,
    file_format: str,
    *,
    original_name: str,
) -> tuple[str, str, dict[str, Any], list[str]]:
    warnings: list[str] = []
    if file_format == "mpp":
        return _extract_mpp_document(path, original_name=original_name)

    metadata = _binary_metadata(path, file_format, warnings)
    limit_warning = _limit_warning(file_format, metadata)
    if limit_warning:
        warnings.append(limit_warning)
        return "", _title_from_name(original_name), metadata, warnings

    markdown = _markitdown_convert(path, warnings)
    title = _markdown_title(markdown) or _title_from_name(original_name)
    plain_text = _markdown_to_plain_text(markdown)

    if file_format == "pdf":
        warnings.append("PDF extraction is text-only; OCR is disabled.")
    elif file_format == "docx":
        warnings.append(
            "DOCX extraction keeps text content only; "
            "layout/comments/track changes are not preserved."
        )
    elif file_format == "xlsx":
        warnings.append(
            "XLSX extraction keeps sheet text and values only; "
            "layout, charts, and formulas are not preserved faithfully."
        )
    elif file_format == "pptx":
        warnings.append(
            "PPTX extraction keeps slide text only; "
            "layout and visual fidelity are not preserved."
        )

    return plain_text, title, metadata, warnings


def _extract_mpp_document(
    path: Path,
    *,
    original_name: str,
) -> tuple[str, str, dict[str, Any], list[str]]:
    warnings: list[str] = []
    payload = _run_mpp_reader(path, warnings)
    summary = payload.get("summary", {}) if isinstance(payload, dict) else {}
    if not isinstance(summary, dict):
        summary = {}
    tasks_raw = payload.get("tasks", []) if isinstance(payload, dict) else []
    if not isinstance(tasks_raw, list):
        tasks_raw = []
    tasks = [task for task in tasks_raw if isinstance(task, dict)]

    title = str(summary.get("file_name") or "").strip() or _title_from_name(
        original_name
    )
    metadata = {
        key: value
        for key, value in summary.items()
        if key != "file_name" and value not in (None, "", [], {})
    }
    metadata["tasks_extracted"] = len(tasks)

    plain_text = _render_mpp_plain_text(title=title, summary=summary, tasks=tasks)
    warnings.append(
        "MPP extraction keeps project/task text only; "
        "timeline, Gantt layout, and native formatting are not preserved."
    )
    return plain_text, title, metadata, warnings


def _run_mpp_reader(path: Path, warnings: list[str]) -> dict[str, Any]:
    script_path = next(
        (
            candidate
            for candidate in _MPP_READER_SCRIPT_CANDIDATES
            if candidate.exists()
        ),
        None,
    )
    if script_path is None:
        warnings.append("MPP reader skill is unavailable.")
        return {}

    command = [sys.executable, str(script_path), str(path)]
    try:
        proc = subprocess.run(
            command,
            capture_output=True,
            text=True,
            check=False,
            timeout=MPP_READER_TIMEOUT_SECONDS,
        )
    except subprocess.TimeoutExpired:
        warnings.append(
            f"MPP reader timed out after {MPP_READER_TIMEOUT_SECONDS}s."
        )
        return {}

    if proc.returncode != 0:
        message = (
            proc.stderr.strip()
            or proc.stdout.strip()
            or "unknown mpp reader error"
        )
        warnings.append(f"MPP reader failed: {message}")
        return {}

    try:
        payload = extract_first_json_dict(
            proc.stdout,
            error_context="mpp reader output",
        )
    except ValueError:
        warnings.append("MPP reader returned invalid JSON.")
        return {}
    return payload if isinstance(payload, dict) else {}


def _render_mpp_plain_text(
    *,
    title: str,
    summary: dict[str, Any],
    tasks: list[dict[str, Any]],
) -> str:
    lines = [title, "", "Project summary:"]
    summary_fields = (
        ("start_date", "Start"),
        ("finish_date", "Finish"),
        ("total_tasks", "Total tasks"),
        ("milestones", "Milestones"),
        ("critical_tasks", "Critical tasks"),
        ("resources_count", "Resources"),
        ("percent_complete", "Percent complete"),
    )
    for key, label in summary_fields:
        value = summary.get(key)
        if value in (None, "", [], {}):
            continue
        suffix = "%" if key == "percent_complete" else ""
        lines.append(f"- {label}: {value}{suffix}")

    if tasks:
        lines.extend(["", "Tasks:"])
    for index, task in enumerate(tasks, start=1):
        name = str(task.get("name") or f"Task {index}").strip()
        lines.append(f"{index}. {name}")
        task_fields = (
            ("id", "ID"),
            ("wbs", "WBS"),
            ("duration", "Duration"),
            ("start", "Start"),
            ("finish", "Finish"),
            ("percent_complete", "Percent complete"),
            ("priority", "Priority"),
            ("resource_names", "Resources"),
            ("predecessors", "Predecessors"),
            ("notes", "Notes"),
            ("milestone", "Milestone"),
            ("critical", "Critical"),
            ("baseline_start", "Baseline start"),
            ("baseline_finish", "Baseline finish"),
            ("actual_start", "Actual start"),
            ("actual_finish", "Actual finish"),
            ("constraint_type", "Constraint type"),
            ("constraint_date", "Constraint date"),
        )
        for key, label in task_fields:
            value = task.get(key)
            if value in (None, "", [], {}):
                continue
            suffix = "%" if key == "percent_complete" else ""
            lines.append(f"   {label}: {value}{suffix}")

    return _normalize_plain_text("\n".join(lines))


def _markitdown_convert(path: Path, warnings: list[str]) -> str:
    try:
        from markitdown import MarkItDown
    except Exception as exc:  # pragma: no cover - guarded by runtime dependency
        warnings.append(f"MarkItDown is unavailable: {exc}")
        return ""

    try:
        converter = MarkItDown(enable_plugins=False)
        result = converter.convert(str(path))
    except Exception as exc:
        warnings.append(f"Converter failed: {exc}")
        return ""

    content = getattr(result, "text_content", "") or getattr(result, "markdown", "")
    return str(content or "").strip()


def _binary_metadata(
    path: Path,
    file_format: str,
    warnings: list[str],
) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    try:
        if file_format == "pdf":
            import pypdfium2 as pdfium

            metadata["pages"] = len(pdfium.PdfDocument(str(path)))
        elif file_format == "docx":
            with ZipFile(path) as archive:
                document_xml = archive.read("word/document.xml")
            root = ET.fromstring(document_xml)
            namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            metadata["paragraphs"] = len(root.findall(".//w:p", namespace))
            metadata["tables"] = len(root.findall(".//w:tbl", namespace))
        elif file_format == "xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                metadata["sheets"] = len(workbook.sheetnames)
                metadata["sheet_names"] = list(workbook.sheetnames)
            finally:
                workbook.close()
        elif file_format == "pptx":
            from pptx import Presentation

            presentation = Presentation(str(path))
            metadata["slides"] = len(presentation.slides)
    except Exception as exc:
        warnings.append(f"Metadata extraction failed: {exc}")
    return metadata


def _limit_warning(file_format: str, metadata: dict[str, Any]) -> str:
    pages = metadata.get("pages")
    slides = metadata.get("slides")
    sheets = metadata.get("sheets")
    if (
        file_format == "pdf"
        and isinstance(pages, int)
        and pages > DOCUMENT_MAX_PDF_PAGES
    ):
        return (
            f"PDF page limit exceeded ({pages} > {DOCUMENT_MAX_PDF_PAGES}); "
            "text extraction was skipped."
        )
    if (
        file_format == "pptx"
        and isinstance(slides, int)
        and slides > DOCUMENT_MAX_PPTX_SLIDES
    ):
        return (
            f"PPTX slide limit exceeded ({slides} > {DOCUMENT_MAX_PPTX_SLIDES}); "
            "text extraction was skipped."
        )
    if (
        file_format == "xlsx"
        and isinstance(sheets, int)
        and sheets > DOCUMENT_MAX_XLSX_SHEETS
    ):
        return (
            f"XLSX sheet limit exceeded ({sheets} > {DOCUMENT_MAX_XLSX_SHEETS}); "
            "text extraction was skipped."
        )
    return ""


def _decode_text_bytes(data: bytes) -> tuple[str, str, list[str]]:
    warnings: list[str] = []
    has_utf16_bom = data.startswith((b"\xff\xfe", b"\xfe\xff"))
    null_ratio = data.count(b"\x00") / max(len(data), 1)
    if has_utf16_bom or null_ratio > 0.2:
        encodings = ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "cp1251")
    else:
        encodings = ("utf-8-sig", "cp1251", "utf-16", "utf-16-le", "utf-16-be")
    for encoding in encodings:
        try:
            decoded = data.decode(encoding)
            if encoding != "utf-8-sig":
                warnings.append(f"Decoded text using fallback encoding {encoding}.")
            return decoded, encoding, warnings
        except UnicodeDecodeError:
            continue

    decoded = data.decode("utf-8", errors="replace")
    warnings.append("Decoded text with replacement characters after encoding fallback.")
    return decoded, "utf-8-replace", warnings


def _html_to_text(raw_html: str) -> dict[str, str]:
    title_match = re.search(
        r"<title[^>]*>(.*?)</title>",
        raw_html,
        re.IGNORECASE | re.DOTALL,
    )
    title = html.unescape(title_match.group(1)).strip() if title_match else ""
    text = re.sub(
        r"<(script|style|noscript)[^>]*>.*?</\1>",
        " ",
        raw_html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    return {"title": title, "content": text}


def _markdown_title(text: str) -> str:
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            candidate = stripped.lstrip("#").strip()
            if candidate:
                return candidate
    for line in text.splitlines():
        stripped = line.strip()
        if stripped:
            return stripped
    return ""


def _markdown_to_plain_text(text: str) -> str:
    content = text.replace("\r\n", "\n").replace("\r", "\n")
    content = re.sub(r"^---\n.*?\n---\n*", "", content, flags=re.DOTALL)
    content = re.sub(r"!\[([^\]]*)\]\([^)]+\)", r"\1", content)
    content = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", content)
    content = re.sub(r"`{1,3}", "", content)
    content = re.sub(r"^\s{0,3}#{1,6}\s*", "", content, flags=re.MULTILINE)
    content = re.sub(r"^\s{0,3}>\s*", "", content, flags=re.MULTILINE)
    content = re.sub(r"^\s*[-*+]\s+", "", content, flags=re.MULTILINE)
    content = re.sub(r"^\s*\d+\.\s+", "", content, flags=re.MULTILINE)
    content = re.sub(r"[*_~|]+", " ", content)
    return _normalize_plain_text(content)


def _normalize_plain_text(text: str) -> str:
    content = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in content.splitlines()]
    cleaned = "\n".join(line for line in lines if line)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    return cleaned.strip()


def _title_from_name(original_name: str) -> str:
    stem = Path(original_name or "document").stem
    title = re.sub(r"[_-]+", " ", stem).strip()
    return title or "Document"


def _unique_ordered(items: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for item in items:
        cleaned = item.strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        ordered.append(cleaned)
    return ordered
